import os
from openpyxl import Workbook
import pandas as pd

f_path = 'file/TaiZhang.xlsx'

read_f = pd.read_excel(f_path)

set_column = ['序号', '设备编号', '设备名称', '设备分类', '工艺段',
                   '设备型号', '主要设备参数', '电机台/功率',
                   '出厂时间/编号', '安装位置', '生产厂家名称',
                   '供应商名称', '启用年月', '原值', '折旧年限', '备注']
read_f.columns = set_column

template = read_f[['序号','设备编号']].copy()

template.head()
print(template)

# wb = Workbook()
# ws = wb.active
#
# record = {
#     '设备名' :'test'
# }
#
# ws.merge_cells('A2:D2')
# ws['A2'] = record['设备名']
#
# wb.save('test.xlsx')
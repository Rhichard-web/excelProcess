import pandas as pd
import os
import shutil
from openpyxl import Workbook
from merge_cell_template import merge_cells,set_cells
from openpyxl.styles import Alignment,Border,Side
from openpyxl.utils import get_column_letter
import time

# Load the "设备台账记录表（麻涌二期）" file
file_path_TZ = 'file/TaiZhang.xlsx'
file_path_template = 'file/template.xlsx'
data_formate_output_path = 'output/data/dataALL'
# muban_df = pd.read_excel(file_path_muban)

# Display the first few rows of the dataframe
# TZ_df.head()



def reFormat(reFor):
    # Data cleaning for "设备台账记录表（麻涌二期）"
    start_time = time.time()
    # Rename columns based on the content
    zhongqi_columns = ['序号', '设备编号', '设备名称', '设备分类', '工艺段',
                       '设备型号', '主要设备参数', '电机台/功率',
                       '出厂时间/编号', '安装位置', '生产厂家名称',
                       '供应商名称', '启用年月', '原值', '折旧年限', '备注']

    # Updating column names
    reFor.columns = zhongqi_columns

    # Remove unnecessary newlines and whitespace
    for col in zhongqi_columns:
        reFor[col] = reFor[col].astype(str).str.replace('\n', '').replace(' ', ',')

    # Dropping the first row which was previously headers
    reFor.drop(0, inplace=True)

    # Display cleaned data
    reFor.head()

    # Selecting and reorganizing columns to match the format of "1提升泵"

    # Extracting the relevant columns
    template_df = reFor[['序号','设备编号','设备名称', '设备型号', '出厂时间/编号', '主要设备参数','生产厂家名称', '电机台/功率','启用年月','原值',]].copy()

    # Renaming columns to match "1提升泵" format
    template_df.rename(columns={
        '设备型号': '型号',
        '出厂时间/编号': '出厂编号',
        '主要设备参数': '主要参数',
        '生产厂家名称' : '厂家名称',
        '电机台/功率': '电机台/功率',
        '原值' : '设备原值',
    }, inplace=True)

    # Adding a placeholder for '额定电流' column as it's not directly available
    template_df['额定电流'] = ''  # Placeholder value

    # Reordering columns to match the target format
    template_format = template_df[['设备名称', '型号', '出厂编号', '主要参数', '额定电流', '电机台/功率']]

    # Display the transformed dataframe
    template_format.head()
    # Save the transformed data to a new Excel file
    # TZ_output_file_path = 'output/TZdata.xlsx'
    # template_format.to_excel(TZ_output_file_path, index=False)
    # TZ_output_file_path

    #输出
    # Creating an individual Excel file for each device record
    for index, record in template_df.iterrows():
        name = record['设备名称'].replace("/","_").replace(" ","")
        file_name = f"{record['序号']}_{name}.xlsx"
        file_path = os.path.join(data_formate_output_path, file_name).replace("\\","/")
        create_individual_excel(record, template_df, file_path)

    # Zipping the directory with individual Excel files
    zip_file_path = 'outpu/data/allData.zip'
    shutil.make_archive(data_formate_output_path, 'zip', data_formate_output_path)

    zip_file_path
    end_time = time.time()
    times = end_time - start_time
    print(f"处理完成！用时：{times}s")


    # Function to create an Excel file for each device record
def create_individual_excel(record, template_format, file_path):
    wb = Workbook()
    ws = wb.active
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_aligned_text = Alignment(horizontal='center', vertical='center')
    # 假设我们有一个有数据的工作表，并且想给所有单元格加边框
    # 这里我们假设工作表的大小是从A1到D10，您可以根据实际情况调整范围
    for row in ws.iter_rows(min_row=1, max_row=31, min_col=1, max_col=8):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_aligned_text

    # # Copy the template format
    # for row in template_format.itertuples(index=False):
    #     ws.append(row)
    merge_cells(ws)

    # Insert device record data
    set_cells(ws,record)
    auto_adjust_column_width(ws)
    # Save to a new Excel file
    wb.save(file_path)

def auto_adjust_column_width(ws, max_width=82):
    column_widths = [12, 20, 15, 5.28, 8, 4.5, 8.5, 11.5]
    row_hights = [13,16.4,35,16.4,35,16.2,23.3,23.3,16.2,16.2]
    # 将列宽转换为Excel的列宽单位（假设每个单位大约2.14毫米）
    # excel_column_widths = [width / 2.14 for width in column_widths]
    for i,height in enumerate(row_hights, start=1):
        ws.row_dimensions[i].height = height
    # 设置每列的宽度
    for i, col_width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = col_width
    # for row in ws.iter_rows():
    #     for cell in row:
    #         if cell.value:
    #             # 对每个单元格的文本内容每4个字符后添加换行符：
    #             # wrapped_text = '\n'.join([cell.value[i:i+9] for i in range(0, len(cell.value), 9)])
    #             # cell.value = wrapped_text
    #             # cell.alignment = Alignment(wrapText=True)
    #             text_length = len(str(cell.value))
    #             # 获取列的字母表示（例如，'A'）
    #             col_letter = get_column_letter(cell.column)
    #             # 更新字典中的最大长度
    #             column_widths[col_letter] = max(column_widths.get(col_letter, 0), text_length)

    # 计算总宽度，并进行缩放以适应A4纸张宽度




    # 设置列宽
    # for col_letter, width in column_widths.items():
    #     adjusted_width = (width * scale_factor) * 1.2  # 1.2为宽度修正系数
    #     ws.column_dimensions[col_letter].width = adjusted_width
    #     # 计算总宽度，并进行缩放以适应A4纸张宽度
    #     total_width = sum(column_widths.values())
    #     scale_factor = 1 if total_width == 0 else min(max_width / total_width, 1)



if __name__ == '__main__':
    TZ_df = pd.read_excel(file_path_TZ)
    reFormat(TZ_df)


